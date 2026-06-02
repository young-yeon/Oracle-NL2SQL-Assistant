from io import BytesIO

from openpyxl import Workbook

from app.services.metadata import MetadataImporter


def workbook_bytes() -> bytes:
    workbook = Workbook()
    tables = workbook.active
    tables.title = "tables"
    tables.append(["name", "schema", "display_name", "description", "synonyms"])
    tables.append(["SALES_FACT", "APP", "Sales", "Sales fact", "sales,revenue"])

    columns = workbook.create_sheet("columns")
    columns.append(["table", "name", "data_type", "display_name", "description", "synonyms"])
    columns.append(["SALES_FACT", "AMOUNT", "NUMBER", "Amount", "Sales amount", "total"])

    workbook.create_sheet("relationships").append(
        ["left_table", "left_column", "right_table", "right_column", "join_type", "description"]
    )
    workbook.create_sheet("terms").append(["term", "canonical_type", "canonical_value", "synonyms", "description"])
    metrics = workbook.create_sheet("metrics")
    metrics.append(["name", "expression", "table", "grain", "synonyms", "description"])
    metrics.append(["total_revenue", "SUM(AMOUNT)", "SALES_FACT", "all", "revenue", "Total revenue"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_import_standard_workbook():
    result = MetadataImporter().from_excel(workbook_bytes(), "metadata.xlsx")

    assert not result.errors
    assert len(result.catalog.tables) == 1
    assert len(result.catalog.columns) == 1
    assert len(result.catalog.metrics) == 1
    assert result.catalog.tables[0].qualified_name == "APP.SALES_FACT"

