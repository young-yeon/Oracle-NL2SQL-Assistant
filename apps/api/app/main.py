from __future__ import annotations

from io import BytesIO

from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook

from app.config import get_settings
from app.schemas import (
    CatalogResponse,
    ChatRequest,
    ChatResponse,
    HealthResponse,
    OracleConnectionTestResponse,
    OracleSettingsPayload,
    OracleSettingsResponse,
    SqlExecuteRequest,
    SqlPreviewRequest,
    UploadResponse,
)
from app.services.audit import AuditLogger
from app.services.guardrails import GuardrailsService
from app.services.llm import LLMService
from app.services.metadata import CatalogStore, MetadataImporter
from app.services.oracle import OracleClient
from app.services.pipeline import QueryPipeline
from app.services.runtime_settings import RuntimeSettingsStore

settings = get_settings()
settings.storage_dir.mkdir(parents=True, exist_ok=True)
runtime_settings = RuntimeSettingsStore(settings.storage_dir)
runtime_settings.apply(settings)

catalog_store = CatalogStore(settings.storage_dir)
guardrails = GuardrailsService(settings)
llm = LLMService(settings)
oracle = OracleClient(settings)
audit = AuditLogger(settings.storage_dir)
pipeline = QueryPipeline(settings, catalog_store, guardrails, llm, oracle, audit)
importer = MetadataImporter()

app = FastAPI(title="Oracle NL2SQL Assistant", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def require_admin_access(x_setup_token: str | None = Header(default=None, alias="X-Setup-Token")) -> None:
    if settings.admin_setup_token and x_setup_token != settings.admin_setup_token:
        raise HTTPException(status_code=401, detail="Admin setup token is required.")


def clean_optional(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def apply_oracle_payload(payload: OracleSettingsPayload, target_settings) -> None:
    target_settings.oracle_dsn = clean_optional(payload.dsn)
    target_settings.oracle_user = clean_optional(payload.user)
    if payload.clear_password:
        target_settings.oracle_password = None
    elif payload.password is not None and payload.password.strip():
        target_settings.oracle_password = payload.password.strip()
    target_settings.oracle_current_schema = clean_optional(payload.current_schema)
    target_settings.oracle_mode = payload.mode
    target_settings.oracle_client_lib_dir = clean_optional(payload.client_lib_dir)
    target_settings.sql_max_rows = payload.sql_max_rows
    target_settings.sql_timeout_seconds = payload.sql_timeout_seconds


def oracle_settings_response(source_settings) -> OracleSettingsResponse:
    mode = source_settings.oracle_mode.lower()
    return OracleSettingsResponse(
        dsn=clean_optional(source_settings.oracle_dsn),
        user=clean_optional(source_settings.oracle_user),
        password_set=bool(source_settings.oracle_password),
        current_schema=clean_optional(source_settings.oracle_current_schema),
        mode=mode if mode in {"thin", "thick"} else "thin",
        client_lib_dir=clean_optional(source_settings.oracle_client_lib_dir),
        sql_max_rows=source_settings.sql_max_rows,
        sql_timeout_seconds=source_settings.sql_timeout_seconds,
        configured=source_settings.oracle_configured,
    )


@app.get("/api/health", response_model=HealthResponse)
async def health() -> HealthResponse:
    version = catalog_store.current_version()
    return HealthResponse(
        status="ok",
        llm_configured=settings.llm_configured,
        nemo_guardrails_enabled=settings.nemo_guardrails_enabled,
        nemo_guardrails_ready=guardrails.ready,
        oracle_configured=settings.oracle_configured,
        oracle_reachable=await oracle.healthcheck(),
        metadata_loaded=bool(version),
        metadata_version=version,
    )


@app.get("/api/settings/oracle", response_model=OracleSettingsResponse)
async def get_oracle_settings(_: None = Depends(require_admin_access)) -> OracleSettingsResponse:
    return oracle_settings_response(settings)


@app.post("/api/settings/oracle", response_model=OracleSettingsResponse)
async def save_oracle_settings(
    payload: OracleSettingsPayload,
    _: None = Depends(require_admin_access),
) -> OracleSettingsResponse:
    apply_oracle_payload(payload, settings)
    runtime_settings.save_oracle(settings)
    oracle.reset()
    pipeline.refresh_settings()
    return oracle_settings_response(settings)


@app.post("/api/settings/oracle/test", response_model=OracleConnectionTestResponse)
async def test_oracle_settings(
    payload: OracleSettingsPayload,
    _: None = Depends(require_admin_access),
) -> OracleConnectionTestResponse:
    candidate_settings = settings.model_copy()
    apply_oracle_payload(payload, candidate_settings)
    candidate = OracleClient(candidate_settings)
    reachable = await candidate.healthcheck()
    if reachable is True:
        message = "Oracle connection succeeded."
    elif reachable is False:
        message = "Oracle connection failed."
    else:
        message = "Oracle connection is not configured."
    return OracleConnectionTestResponse(
        configured=candidate_settings.oracle_configured,
        reachable=reachable,
        message=message,
    )


@app.post("/api/chat", response_model=ChatResponse)
async def chat(request: ChatRequest) -> ChatResponse:
    return await pipeline.chat(request.session_id, request.message, request.metadata_version)


@app.post("/api/sql/preview", response_model=ChatResponse)
async def preview(request: SqlPreviewRequest) -> ChatResponse:
    return await pipeline.preview(request.message, request.metadata_version)


@app.post("/api/sql/execute", response_model=ChatResponse)
async def execute_sql(request: SqlExecuteRequest) -> ChatResponse:
    return await pipeline.execute_approved_sql(
        request.session_id,
        request.message,
        request.sql,
        request.metadata_version,
    )


@app.post("/api/metadata/upload", response_model=UploadResponse)
async def upload_metadata(file: UploadFile = File(...)) -> UploadResponse:
    content = await file.read()
    result = importer.from_excel(content, filename=file.filename or "")
    if not result.errors:
        catalog_store.save(result.catalog)
    return UploadResponse(
        version=result.catalog.version,
        tables=len(result.catalog.tables),
        columns=len(result.catalog.columns),
        relationships=len(result.catalog.relationships),
        terms=len(result.catalog.terms),
        metrics=len(result.catalog.metrics),
        warnings=result.warnings,
        errors=result.errors,
    )


@app.get("/api/metadata/catalog", response_model=CatalogResponse)
async def catalog(version: str | None = None) -> CatalogResponse:
    loaded = catalog_store.load(version)
    if not loaded:
        return CatalogResponse(warnings=["No metadata catalog is loaded."])
    return CatalogResponse(**loaded.summary())


@app.get("/api/metadata/template")
async def metadata_template() -> StreamingResponse:
    workbook = Workbook()
    default = workbook.active
    default.title = "tables"
    default.append(["name", "schema", "display_name", "description", "synonyms"])
    default.append(["SALES_FACT", "APP", "Sales", "Sales transaction fact table", "orders,revenue"])

    columns = workbook.create_sheet("columns")
    columns.append(["table", "name", "data_type", "display_name", "description", "synonyms", "semantic_type", "nullable"])
    columns.append(["SALES_FACT", "SALE_DATE", "DATE", "Sale date", "Transaction date", "order date", "date", "N"])
    columns.append(["SALES_FACT", "AMOUNT", "NUMBER", "Amount", "Sales amount", "revenue,total", "amount", "Y"])

    relationships = workbook.create_sheet("relationships")
    relationships.append(["left_table", "left_column", "right_table", "right_column", "join_type", "description"])

    terms = workbook.create_sheet("terms")
    terms.append(["term", "canonical_type", "canonical_value", "synonyms", "description"])
    terms.append(["revenue", "metric", "total_revenue", "sales,total sales", "Total sales amount"])

    metrics = workbook.create_sheet("metrics")
    metrics.append(["name", "expression", "table", "grain", "synonyms", "description"])
    metrics.append(["total_revenue", "SUM(AMOUNT)", "SALES_FACT", "all", "revenue,sales", "Total sales amount"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": "attachment; filename=metadata_template.xlsx"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


if settings.web_dist_dir:
    web_dist_dir = settings.web_dist_dir
    index_file = web_dist_dir / "index.html"
    assets_dir = web_dist_dir / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=assets_dir), name="web-assets")

    if index_file.exists():

        @app.get("/")
        async def web_index() -> FileResponse:
            return FileResponse(index_file)

        @app.get("/{full_path:path}")
        async def web_spa_fallback(full_path: str) -> FileResponse:
            if full_path.startswith("api/"):
                raise HTTPException(status_code=404, detail="Not found")
            return FileResponse(index_file)
